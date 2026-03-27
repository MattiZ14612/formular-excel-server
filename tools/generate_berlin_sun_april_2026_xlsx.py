import openpyxl
from openpyxl.utils import get_column_letter

# Dates and times for April 2026 in Berlin (CEST)
dates = [
    "2026-04-01", "2026-04-02", "2026-04-03", "2026-04-04", "2026-04-05",
    "2026-04-06", "2026-04-07", "2026-04-08", "2026-04-09", "2026-04-10",
    "2026-04-11", "2026-04-12", "2026-04-13", "2026-04-14", "2026-04-15",
    "2026-04-16", "2026-04-17", "2026-04-18", "2026-04-19", "2026-04-20",
    "2026-04-21", "2026-04-22", "2026-04-23", "2026-04-24", "2026-04-25",
    "2026-04-26", "2026-04-27", "2026-04-28", "2026-04-29", "2026-04-30"
]
sunrise_times = [
    "06:04", "06:03", "06:01", "05:59", "05:58",
    "05:56", "05:55", "05:53", "05:52", "05:50",
    "05:49", "05:47", "05:46", "05:44", "05:43",
    "05:41", "05:40", "05:38", "05:37", "05:35",
    "05:34", "05:32", "05:31", "05:29", "05:28",
    "05:26", "05:25", "05:23", "05:22", "05:20"
]
sunset_times = [
    "20:00", "20:01", "20:02", "20:03", "20:05",
    "20:06", "20:07", "20:08", "20:09", "20:10",
    "20:12", "20:13", "20:14", "20:15", "20:16",
    "20:17", "20:18", "20:19", "20:20", "20:21",
    "20:22", "20:23", "20:24", "20:25", "20:26",
    "20:27", "20:28", "20:29", "20:30", "20:31"
]

# Create a new workbook and sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Sonnenaufgang und Sonnenuntergang"

# Header row
headers = ["Datum", "Sonnenaufgang", "Sonnenuntergang", "Tageslänge"]
for col_num, header in enumerate(headers, 1):
    sheet[get_column_letter(col_num) + '1'] = header

# Populate the sheet with the dates and times
for i, date_str in enumerate(dates):
    date_cell = f'A{i + 2}'
    sunrise_cell = f'B{i + 2}'
    sunset_cell = f'C{i + 2}'
    duration_cell = f'D{i + 2}'
    sheet[date_cell] = date_str
    sheet[sunrise_cell] = date_str + ' ' + sunrise_times[i]
    sheet[sunset_cell] = date_str + ' ' + sunset_times[i]
    sheet[duration_cell] = f'={sunset_cell}-{sunrise_cell}'

# Format cells
for i in range(2, 32):
    sheet[f'A{i}'].number_format = 'YYYY-MM-DD'
    sheet[f'B{i}'].number_format = 'HH:MM'
    sheet[f'C{i}'].number_format = 'HH:MM'
    sheet[f'D{i}'].number_format = '[h]:mm'

# Save the workbook
wb.save("berlin_sun_april_2026.xlsx")